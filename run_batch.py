# run_batch.py
# Runs the mapping pipeline on all 101 base controls from the official CIS-to-NIST
# ground truth mapping. Saves JSON, PDF, and a running summary spreadsheet to full_runs/.
#
# Usage:
#   python run_batch.py
#
# Crash-safe: each control is saved immediately after it completes.
# If the run is interrupted, already-completed controls are skipped on resume.

import json
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agents import nist_controls, run_mapping_pipeline, FULL_RUNS_PATH
from generate_pdf import generate_report
from parse_results import parse_classifications, extract_coverage

# ============================================================
# 101 base controls from the official CIS-NIST mapping
# NOTE: Set to 3 controls for testing -- uncomment full list when ready
# ============================================================

# PHASE1_CONTROLS = [
#     "AC-2", "AU-11", "CM-7",
# ]

# FULL LIST -- 101 base controls from the official CIS-NIST ground truth mapping
PHASE1_CONTROLS = [
    # AC - Access Control (15)
    "AC-1", "AC-2", "AC-3", "AC-4", "AC-5", "AC-6", "AC-7",
    "AC-11", "AC-12", "AC-17", "AC-18", "AC-19", "AC-20", "AC-21", "AC-22",
    # AT - Awareness and Training (3)
    "AT-1", "AT-2", "AT-3",
    # AU - Audit and Accountability (9)
    "AU-1", "AU-2", "AU-3", "AU-4", "AU-6", "AU-7", "AU-8", "AU-11", "AU-12",
    # CA - Assessment, Authorization (3)
    "CA-5", "CA-7", "CA-9",
    # CM - Configuration Management (9)
    "CM-1", "CM-2", "CM-6", "CM-7", "CM-8", "CM-9", "CM-10", "CM-11", "CM-12",
    # CP - Contingency Planning (7)
    "CP-2", "CP-4", "CP-6", "CP-7", "CP-8", "CP-9", "CP-10",
    # IA - Identification and Authentication (2)
    "IA-4", "IA-5",
    # IR - Incident Response (7)
    "IR-1", "IR-3", "IR-4", "IR-5", "IR-6", "IR-7", "IR-8",
    # MA - Maintenance (2)
    "MA-3", "MA-4",
    # MP - Media Protection (4)
    "MP-2", "MP-5", "MP-6", "MP-7",
    # PL - Planning (1)
    "PL-8",
    # PM - Program Management (5)
    "PM-5", "PM-7", "PM-13", "PM-17", "PM-30",
    # RA - Risk Assessment (4)
    "RA-1", "RA-2", "RA-5", "RA-7",
    # SA - System and Services Acquisition (8)
    "SA-3", "SA-4", "SA-8", "SA-9", "SA-10", "SA-11", "SA-15", "SA-22",
    # SC - System and Communications Protection (10)
    "SC-4", "SC-7", "SC-8", "SC-18", "SC-20", "SC-21", "SC-22", "SC-23", "SC-28", "SC-39",
    # SI - System and Information Integrity (7)
    "SI-2", "SI-3", "SI-4", "SI-7", "SI-8", "SI-12", "SI-16",
    # SR - Supply Chain Risk Management (5)
    "SR-1", "SR-5", "SR-6", "SR-11", "SR-12",
]

# ============================================================
# Spreadsheet setup
# ============================================================

XLSX_PATH = os.path.join(FULL_RUNS_PATH, "batch_results.xlsx")

HEADERS = [
    "NIST ID", "Title", "Family",
    "Strong Mappings", "Partial Mappings", "Weak Mappings",
    "Strong Count", "Partial Count", "Weak Count",
    "Coverage %", "Elapsed (s)", "Timestamp", "Status"
]

HEADER_COLOR = "1a1a2e"
HEADER_FONT_COLOR = "FFFFFF"
STRONG_COLOR = "d4edda"
PARTIAL_COLOR = "fff3cd"
WEAK_COLOR = "f8d7da"
ALT_COLOR = "f0f0f8"

COL_WIDTHS = [10, 35, 25, 35, 45, 35, 10, 10, 10, 12, 12, 20, 10]

def create_spreadsheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"

    # header row
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"  Spreadsheet created: {XLSX_PATH}")


def append_to_spreadsheet(result, status="complete"):
    """Parse result dict and append a row to the spreadsheet."""
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        create_spreadsheet()
        wb = load_workbook(XLSX_PATH)
        ws = wb.active

    strong, partial, weak = parse_classifications(result.get("mapping_classification", ""))
    coverage = extract_coverage(result.get("interpretation", ""))

    row_num = ws.max_row + 1
    row_data = [
        result.get("nist_id", ""),
        result.get("nist_title", ""),
        result.get("nist_family", ""),
        ", ".join(strong),
        ", ".join(partial),
        ", ".join(weak),
        len(strong),
        len(partial),
        len(weak),
        coverage,
        result.get("elapsed_seconds", 0),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        status,
    ]

    # alternate row color
    fill_color = ALT_COLOR if row_num % 2 == 0 else "FFFFFF"

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=fill_color)

    ws.row_dimensions[row_num].height = 40
    wb.save(XLSX_PATH)


# ============================================================
# Batch runner
# ============================================================

def get_completed_ids():
    """Return set of NIST IDs already saved in full_runs/ as JSON files."""
    completed = set()
    if os.path.exists(FULL_RUNS_PATH):
        for fname in os.listdir(FULL_RUNS_PATH):
            if fname.endswith(".json"):
                nist_id = fname.split("_pipeline")[0]
                completed.add(nist_id)
    return completed


def run_batch():
    os.makedirs(FULL_RUNS_PATH, exist_ok=True)

    # initialize spreadsheet if it doesn't exist
    if not os.path.exists(XLSX_PATH):
        create_spreadsheet()

    completed = get_completed_ids()
    remaining = [c for c in PHASE1_CONTROLS if c not in completed]

    print(f"\n{'='*60}")
    print(f"PHASE 1 BATCH RUN")
    print(f"{'='*60}")
    print(f"Total controls: {len(PHASE1_CONTROLS)}")
    print(f"Already completed: {len(completed)}")
    print(f"Remaining: {len(remaining)}")
    print(f"Output folder: {FULL_RUNS_PATH}")
    print(f"Spreadsheet: {XLSX_PATH}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    if not remaining:
        print("All controls already completed!")
        return

    batch_start = time.time()
    success_count = 0
    error_count = 0

    for i, control_id in enumerate(remaining, 1):
        print(f"\n[{i}/{len(remaining)}] Processing {control_id}...")

        # find control in nist_controls
        control = next((c for c in nist_controls if c["id"] == control_id), None)
        if control is None:
            print(f"  WARNING: {control_id} not found in nist_controls_parsed.json -- skipping")
            error_count += 1
            continue

        try:
            result = run_mapping_pipeline(control)

            # save JSON
            json_path = os.path.join(FULL_RUNS_PATH, f"{control_id}_pipeline.json")
            with open(json_path, "w") as f:
                json.dump(result, f, indent=2)
            print(f"  JSON saved: {json_path}")

            # save PDF
            pdf_path = os.path.join(FULL_RUNS_PATH, f"{control_id}_pipeline.pdf")
            generate_report(result, pdf_path)

            # append to spreadsheet
            append_to_spreadsheet(result, status="complete")
            print(f"  Spreadsheet updated")

            success_count += 1

        except Exception as e:
            print(f"  ERROR on {control_id}: {e}")
            append_to_spreadsheet(
                {"nist_id": control_id, "nist_title": "", "nist_family": "",
                 "mapping_classification": "", "interpretation": "", "elapsed_seconds": 0},
                status=f"error: {str(e)[:50]}"
            )
            error_count += 1
            continue

        # progress update every 10 controls
        if i % 10 == 0:
            elapsed = round(time.time() - batch_start, 0)
            avg = elapsed / i
            remaining_time = round(avg * (len(remaining) - i) / 60, 0)
            print(f"\n  --- Progress: {i}/{len(remaining)} complete | "
                  f"Elapsed: {elapsed}s | Est. remaining: {remaining_time} min ---\n")

    total_time = round(time.time() - batch_start, 0)
    print(f"\n{'='*60}")
    print(f"BATCH COMPLETE")
    print(f"{'='*60}")
    print(f"Successful: {success_count}")
    print(f"Errors: {error_count}")
    print(f"Total time: {total_time}s ({round(total_time/3600, 1)} hours)")
    print(f"Results saved to: {FULL_RUNS_PATH}")


if __name__ == "__main__":
    run_batch()